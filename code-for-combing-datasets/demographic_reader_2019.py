from openpyxl import load_workbook

# Load spreadsheets
total_wb = load_workbook(filename='fall-membership-total-2019.xlsx')
total_ws = total_wb.active
black_wb = load_workbook(filename='fall-membership-black-2019.xlsx')
black_ws = black_wb.active

# Get bounds of spreadsheets
total_row_count = total_ws.max_row
total_col_count = total_ws.max_column
black_row_count = black_ws.max_row
black_col_count = black_ws.max_column

def get_school_names():
	school_names = []

	for i in range(2, total_row_count+1):

		school = total_ws.cell(row=i, column=5).value.strip()
		
		if school not in school_names:
			school_names.append(school)

	return school_names

def get_total_student_count():
	school_to_count = dict()

	for i in range(2, total_row_count+1):

		school = total_ws.cell(row=i, column=5).value.strip()
		count = total_ws.cell(row=i, column=8).value

		school_to_count[school] = count

	return school_to_count

def get_black_student_count():
	school_to_count = dict()

	for i in range(2, black_row_count+1):

		school = black_ws.cell(row=i, column=5).value.strip()
		count = black_ws.cell(row=i, column=9).value

		school_to_count[school] = count

	return school_to_count