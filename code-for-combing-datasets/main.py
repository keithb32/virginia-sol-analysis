from openpyxl import Workbook
import sol_reader, demographic_reader_2019, mode_reader

SUBJECTS = ["English:Reading", "Mathematics", "Science"]
DEMOGRAPHICS = ["Black", "All Students"]
MODES_OF_INSTRUCTION = ["Inperson, Full-Time", "Remote, Full-Time", "Both Inperson and Remote"]

SCHOOLS = demographic_reader_2019.get_school_names()

BLACK_ENGLISH_2019 = sol_reader.find_pass_rates("English: Reading", "Black", "2019")
BLACK_ENGLISH_2021 = sol_reader.find_pass_rates("English: Reading", "Black", "2021")
BLACK_MATH_2019 = sol_reader.find_pass_rates("Mathematics", "Black", "2019")
BLACK_MATH_2021 = sol_reader.find_pass_rates("Mathematics", "Black", "2021")
BLACK_SCIENCE_2019 = sol_reader.find_pass_rates("Science", "Black", "2019")
BLACK_SCIENCE_2021 = sol_reader.find_pass_rates("Science", "Black", "2021")

ALL_ENGLISH_2019 = sol_reader.find_pass_rates("English: Reading", "All Students", "2019")
ALL_ENGLISH_2021 = sol_reader.find_pass_rates("English: Reading", "All Students", "2021")
ALL_MATH_2019 = sol_reader.find_pass_rates("Mathematics", "All Students", "2019")
ALL_MATH_2021 = sol_reader.find_pass_rates("Mathematics", "All Students", "2021")
ALL_SCIENCE_2019 = sol_reader.find_pass_rates("Science", "All Students", "2019")
ALL_SCIENCE_2021 = sol_reader.find_pass_rates("Science", "All Students", "2021")

TOTAL_STUDENT_COUNT_2019 = demographic_reader_2019.get_total_student_count()
BLACK_STUDENT_COUNT_2019 = demographic_reader_2019.get_black_student_count()
TOTAL_STUDENT_COUNT_2021 = mode_reader.get_total_student_count()
BLACK_STUDENT_COUNT_2021 = mode_reader.get_black_student_count()

BLACK_INPERSON_2021 = mode_reader.get_instruction_mode_count("Inperson, Full-Time", "Black")
BLACK_REMOTE_2021 = mode_reader.get_instruction_mode_count("Remote, Full-Time", "Black")
BLACK_HYBRID_2021 = mode_reader.get_instruction_mode_count("Both Inperson and Remote", "Black")

DATA_MAPS = [BLACK_ENGLISH_2019, BLACK_ENGLISH_2021, BLACK_MATH_2019, BLACK_MATH_2021, BLACK_SCIENCE_2019, BLACK_SCIENCE_2021,
TOTAL_STUDENT_COUNT_2019, BLACK_STUDENT_COUNT_2019, TOTAL_STUDENT_COUNT_2021, BLACK_STUDENT_COUNT_2019, BLACK_INPERSON_2021,
BLACK_REMOTE_2021, BLACK_HYBRID_2021]

# Open workbook to write to
new_wb = Workbook()
new_ws = new_wb.active

# Write headers
new_ws.cell(row=1, column=1).value = 'School-Name'
new_ws.cell(row=1, column=2).value = 'Total-Student-Count-2018-19'
new_ws.cell(row=1, column=3).value = 'Black-Student-Count-2018-19'
new_ws.cell(row=1, column=4).value = 'Black-English-SOL-2018-19'
new_ws.cell(row=1, column=5).value = 'Black-Math-SOL-2018-19'
new_ws.cell(row=1, column=6).value = 'Black-Science-SOL-2018-19'
new_ws.cell(row=1, column=7).value = 'Total-Student-Count-2020-21'
new_ws.cell(row=1, column=8).value = 'Black-Student-Count-2020-21'
new_ws.cell(row=1, column=9).value = 'Black-Inperson-Count-2020-21'
new_ws.cell(row=1, column=10).value = 'Black-Hybrid-Count-2020-21'
new_ws.cell(row=1, column=11).value = 'Black-Remote-Count-2020-21'
new_ws.cell(row=1, column=12).value = 'Black-English-SOL-2020-21'
new_ws.cell(row=1, column=13).value = 'Black-Math-SOL-2020-21'
new_ws.cell(row=1, column=14).value = 'Black-Science-SOL-2020-21'
new_ws.cell(row=1, column=15).value = 'All-English-SOL-2018-2019'
new_ws.cell(row=1, column=16).value = 'All-Math-SOL-2018-2019'
new_ws.cell(row=1, column=17).value = 'All-Science-SOL-2018-2019'
new_ws.cell(row=1, column=18).value = 'All-English-SOL-2020-2021'
new_ws.cell(row=1, column=19).value = 'All-Math-SOL-2020-2021'
new_ws.cell(row=1, column=20).value = 'All-Science-SOL-2020-2021'



for i in range(len(SCHOOLS)):
	school = SCHOOLS[i]
	
	new_ws.cell(row=i+2, column=1).value = school
	
	if school in TOTAL_STUDENT_COUNT_2019:
		new_ws.cell(row=i+2, column=2).value = TOTAL_STUDENT_COUNT_2019[school]
	#else
	if school in BLACK_STUDENT_COUNT_2019:
		new_ws.cell(row=i+2, column=3).value = BLACK_STUDENT_COUNT_2019[school]

	if school in BLACK_ENGLISH_2019:
		new_ws.cell(row=i+2, column=4).value = BLACK_ENGLISH_2019[school] 
	if school in BLACK_MATH_2019:
		new_ws.cell(row=i+2, column=5).value = BLACK_MATH_2019[school]
	if school in BLACK_SCIENCE_2019:
		new_ws.cell(row=i+2, column=6).value = BLACK_SCIENCE_2019[school]

	if school in TOTAL_STUDENT_COUNT_2021:
		new_ws.cell(row=i+2, column=7).value = TOTAL_STUDENT_COUNT_2021[school]
	if school in BLACK_STUDENT_COUNT_2021:
		new_ws.cell(row=i+2, column=8).value = BLACK_STUDENT_COUNT_2021[school]

	if school in BLACK_INPERSON_2021:
		new_ws.cell(row=i+2, column=9).value = BLACK_INPERSON_2021[school]
	if school in BLACK_HYBRID_2021:
		new_ws.cell(row=i+2, column=10).value = BLACK_HYBRID_2021[school]
	if school in BLACK_REMOTE_2021:
		new_ws.cell(row=i+2, column=11).value = BLACK_REMOTE_2021[school]

	if school in BLACK_ENGLISH_2021:
		new_ws.cell(row=i+2, column=12).value = BLACK_ENGLISH_2021[school]
	if school in BLACK_MATH_2021:
		new_ws.cell(row=i+2, column=13).value = BLACK_MATH_2021[school]
	if school in BLACK_SCIENCE_2021:
		new_ws.cell(row=i+2, column=14).value = BLACK_SCIENCE_2021[school]

	if school in ALL_ENGLISH_2019:
		new_ws.cell(row=i+2, column=15).value = ALL_ENGLISH_2019[school]

	if school in ALL_MATH_2019:
		new_ws.cell(row=i+2, column=16).value = ALL_MATH_2019[school]

	if school in ALL_SCIENCE_2019:
		new_ws.cell(row=i+2, column=17).value = ALL_SCIENCE_2019[school]

	if school in ALL_ENGLISH_2021:
		new_ws.cell(row=i+2, column=18).value = ALL_ENGLISH_2021[school]

	if school in ALL_MATH_2021:
		new_ws.cell(row=i+2, column=19).value = ALL_MATH_2021[school]

	if school in ALL_SCIENCE_2021:
		new_ws.cell(row=i+2, column=20).value = ALL_SCIENCE_2021[school]



new_wb.save('pass_rates.xlsx')

