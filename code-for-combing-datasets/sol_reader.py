from openpyxl import load_workbook

# Load spreadsheet
sol_wb = load_workbook(filename='SOL-school.xlsx')
sol_ws = sol_wb.active

# Get bounds of spreadsheet
sol_row_count = sol_ws.max_row
sol_column_count = sol_ws.max_column

def find_pass_rates(desired_subject, desired_demographic, year):
	school_to_passrate = dict()

	if year == '2019':
		pass_rate_col = 11
	elif year == '2021':
		pass_rate_col = 12

	# Loop over every row of the spreadsheet
	for i in range(3, sol_row_count+1):

		# Read values from cells
		school = sol_ws.cell(row=i, column=5).value.strip()
		subject = sol_ws.cell(row=i, column=9).value
		demographic = sol_ws.cell(row=i, column=10).value
		pass_rate = sol_ws.cell(row=i, column=pass_rate_col).value

		# Add school, passrate pair to the dictionary if subject and demographic match what we're looking for
		if subject == desired_subject and demographic == desired_demographic:
			school_to_passrate[school] = pass_rate

	return school_to_passrate








