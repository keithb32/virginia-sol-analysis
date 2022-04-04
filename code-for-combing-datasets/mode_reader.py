from openpyxl import load_workbook

# Load spreadsheet
mode_wb = load_workbook(filename='mode-of-instruction-2021.xlsx')
mode_ws = mode_wb.active

# Get bounds of spreadsheet
mode_row_count = mode_ws.max_row
mode_col_count = mode_ws.max_column

def get_total_student_count():
	school_to_count = dict()

	for i in range(2, mode_row_count+1):
		school = mode_ws.cell(row=i, column=4).value.strip()
		count = mode_ws.cell(row=i, column=6).value

		if count == None or count == '<':
			count = 0

		if school not in school_to_count:
			school_to_count[school] = int(count)
		else:
			school_to_count[school] += int(count)

	return school_to_count

def get_black_student_count():
	school_to_count = dict()

	for i in range(2, mode_row_count+1):
		school = mode_ws.cell(row=i, column=4).value.strip()
		count = mode_ws.cell(row=i, column=10).value

		if count == None or count == '<':
			count = 0

		if school not in school_to_count:
			school_to_count[school] = int(count)
		else:
			school_to_count[school] += int(count)

	return school_to_count

def get_instruction_mode_count(desired_mode, demographic):
	school_to_count = dict()

	if demographic == 'All':
		dem_col = 6
	elif demographic == 'Black':
		dem_col = 10

	for i in range(2, mode_row_count+1):
		school = mode_ws.cell(row=i, column=4).value.strip()
		mode = mode_ws.cell(row=i, column=5).value
		
		if mode == desired_mode:
			count = mode_ws.cell(row=i, column=dem_col).value
			school_to_count[school] = count

	return school_to_count
